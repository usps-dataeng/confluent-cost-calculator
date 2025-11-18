from datetime import datetime
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def generate_cost_projection_csv(
    selected_size,
    partitions,
    storage_gb,
    cku_config,
    flat_costs,
    costs,
    annual_increase_rate=0.034
):
    """Generate CSV format for backward compatibility"""
    current_year = datetime.now().year
    csv_rows = []

    csv_rows.append('Confluent Cloud Cost Calculator - 7 Year Projection')
    csv_rows.append('')
    csv_rows.append(f'T-Shirt Size:,{selected_size}')
    csv_rows.append(f'Partitions:,{partitions}')
    csv_rows.append(f'Storage (GB):,{storage_gb}')
    csv_rows.append(f'Annual Increase Rate:,{annual_increase_rate * 100:.1f}%')
    csv_rows.append('')

    csv_rows.append('CKU Configuration')
    csv_rows.append(f"Azure CKUs:,{cku_config['azure_ckus']}")
    csv_rows.append(f"Azure Rate ($/CKU/Month):,${cku_config['azure_rate']}")
    csv_rows.append(f"GCP CKUs:,{cku_config['gcp_ckus']}")
    csv_rows.append(f"GCP Rate ($/CKU/Month):,${cku_config['gcp_rate']}")
    csv_rows.append('')

    csv_rows.append('Current Year Cost Breakdown')
    csv_rows.append('Category,Annual Cost,Monthly Cost')
    csv_rows.append(f"Compute (CKU),${costs['compute']:.2f},${costs['compute'] / 12:.2f}")
    csv_rows.append(f"Storage,${costs['storage']:.2f},${costs['storage'] / 12:.2f}")
    csv_rows.append(f"Network,${costs['network']:.2f},${costs['network'] / 12:.2f}")
    csv_rows.append(f"Governance,${costs['governance']:.2f},${costs['governance'] / 12:.2f}")
    csv_rows.append(f"Total,${costs['total_yearly']:.2f},${costs['total_monthly']:.2f}")
    csv_rows.append('')

    csv_rows.append('7-Year Cost Projection')
    csv_rows.append('Year,Compute Cost,Storage Cost,Network Cost,Governance Cost,Total Annual Cost,Cumulative Cost')

    cumulative_cost = 0

    for year in range(7):
        year_label = current_year + year
        multiplier = (1 + annual_increase_rate) ** year

        compute_cost = costs['compute'] * multiplier
        storage_cost = costs['storage'] * multiplier
        network_cost = costs['network'] * multiplier
        governance_cost = costs['governance'] * multiplier
        total_cost = compute_cost + storage_cost + network_cost + governance_cost

        cumulative_cost += total_cost

        csv_rows.append(
            f"{year_label},"
            f"${compute_cost:.2f},"
            f"${storage_cost:.2f},"
            f"${network_cost:.2f},"
            f"${governance_cost:.2f},"
            f"${total_cost:.2f},"
            f"${cumulative_cost:.2f}"
        )

    csv_rows.append('')
    csv_rows.append('Monthly Breakdown by Year')
    csv_rows.append('Year,Jan,Feb,Mar,Apr,May,Jun,Jul,Aug,Sep,Oct,Nov,Dec,Annual Total')

    for year in range(7):
        year_label = current_year + year
        multiplier = (1 + annual_increase_rate) ** year
        total_annual = costs['total_yearly'] * multiplier
        monthly_avg = total_annual / 12

        monthly_values = ','.join([f'${monthly_avg:.2f}'] * 12)
        csv_rows.append(f'{year_label},{monthly_values},${total_annual:.2f}')

    return '\n'.join(csv_rows)


def generate_cost_projection_excel(
    selected_size,
    partitions,
    storage_gb,
    cku_config,
    flat_costs,
    costs,
    annual_increase_rate=0.034,
    logo_path='public/Postal Logo.png'
):
    """Generate formatted Excel file with USPS logo"""
    if not HAS_OPENPYXL:
        # Fallback to CSV if openpyxl not available
        return generate_cost_projection_csv(
            selected_size, partitions, storage_gb, cku_config,
            flat_costs, costs, annual_increase_rate
        ).encode('utf-8')

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Projection"

    # Color scheme - USPS Blue
    usps_blue = "004B87"
    header_fill = PatternFill(start_color=usps_blue, end_color=usps_blue, fill_type="solid")
    light_blue = PatternFill(start_color="D9E9F7", end_color="D9E9F7", fill_type="solid")
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    title_font = Font(name='Calibri', size=16, bold=True, color=usps_blue)
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Add USPS logo if available
    try:
        img = Image(logo_path)
        img.width = 200
        img.height = 60
        ws.add_image(img, 'A1')
        start_row = 5
    except:
        start_row = 1

    current_year = datetime.now().year
    row = start_row

    # Title
    ws.merge_cells(f'A{row}:G{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = 'Confluent Cloud Cost Calculator - 7 Year Projection'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Configuration section
    ws[f'A{row}'] = 'T-Shirt Size:'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = selected_size
    row += 1

    ws[f'A{row}'] = 'Partitions:'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = partitions
    row += 1

    ws[f'A{row}'] = 'Storage (GB):'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = storage_gb
    row += 1

    ws[f'A{row}'] = 'Annual Increase Rate:'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = f'{annual_increase_rate * 100:.1f}%'
    row += 2

    # CKU Configuration
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = 'CKU Configuration'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    ws[f'A{row}'] = 'Azure CKUs:'
    ws[f'B{row}'] = cku_config['azure_ckus']
    row += 1
    ws[f'A{row}'] = 'Azure Rate ($/CKU/Month):'
    ws[f'B{row}'] = f"${cku_config['azure_rate']}"
    row += 1
    ws[f'A{row}'] = 'GCP CKUs:'
    ws[f'B{row}'] = cku_config['gcp_ckus']
    row += 1
    ws[f'A{row}'] = 'GCP Rate ($/CKU/Month):'
    ws[f'B{row}'] = f"${cku_config['gcp_rate']}"
    row += 2

    # Current Year Cost Breakdown
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = 'Current Year Cost Breakdown'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    # Headers
    headers = ['Category', 'Annual Cost', 'Monthly Cost']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    # Cost rows
    cost_data = [
        ('Compute (CKU)', costs['compute'], costs['compute'] / 12),
        ('Storage', costs['storage'], costs['storage'] / 12),
        ('Network', costs['network'], costs['network'] / 12),
        ('Governance', costs['governance'], costs['governance'] / 12),
        ('Total', costs['total_yearly'], costs['total_monthly'])
    ]

    for category, annual, monthly in cost_data:
        ws.cell(row, 1, category).border = thin_border
        cell_annual = ws.cell(row, 2, annual)
        cell_annual.number_format = '$#,##0.00'
        cell_annual.border = thin_border
        cell_monthly = ws.cell(row, 3, monthly)
        cell_monthly.number_format = '$#,##0.00'
        cell_monthly.border = thin_border

        if category == 'Total':
            ws.cell(row, 1).font = bold_font
            ws.cell(row, 2).font = bold_font
            ws.cell(row, 3).font = bold_font
            ws.cell(row, 1).fill = light_blue
            ws.cell(row, 2).fill = light_blue
            ws.cell(row, 3).fill = light_blue
        row += 1

    row += 1

    # 7-Year Projection
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = '7-Year Cost Projection'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    projection_headers = ['Year', 'Compute Cost', 'Storage Cost', 'Network Cost',
                          'Governance Cost', 'Total Annual Cost', 'Cumulative Cost']
    for col, header in enumerate(projection_headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    cumulative_cost = 0
    for year in range(7):
        year_label = current_year + year
        multiplier = (1 + annual_increase_rate) ** year

        compute_cost = costs['compute'] * multiplier
        storage_cost = costs['storage'] * multiplier
        network_cost = costs['network'] * multiplier
        governance_cost = costs['governance'] * multiplier
        total_cost = compute_cost + storage_cost + network_cost + governance_cost
        cumulative_cost += total_cost

        ws.cell(row, 1, year_label).border = thin_border
        ws.cell(row, 2, compute_cost).number_format = '$#,##0.00'
        ws.cell(row, 2).border = thin_border
        ws.cell(row, 3, storage_cost).number_format = '$#,##0.00'
        ws.cell(row, 3).border = thin_border
        ws.cell(row, 4, network_cost).number_format = '$#,##0.00'
        ws.cell(row, 4).border = thin_border
        ws.cell(row, 5, governance_cost).number_format = '$#,##0.00'
        ws.cell(row, 5).border = thin_border
        ws.cell(row, 6, total_cost).number_format = '$#,##0.00'
        ws.cell(row, 6).border = thin_border
        ws.cell(row, 7, cumulative_cost).number_format = '$#,##0.00'
        ws.cell(row, 7).border = thin_border

        # Alternate row colors
        if year % 2 == 0:
            for col in range(1, 8):
                ws.cell(row, col).fill = light_blue
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
