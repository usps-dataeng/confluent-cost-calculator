from datetime import datetime
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_technical_export(
    data_volume_gb_day,
    message_rate,
    avg_message_size_kb,
    retention_days,
    partitions,
    replication_factor,
    peak_avg_ratio,
    costs,
    storage_cost_per_gb_month,
    throughput_cost_per_mbps_month,
    network_cost_per_gb_month,
    partition_cost_per_partition_month,
    retention_overhead_pct
):
    """
    Generate comprehensive technical cost model export matching the React version format.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for technical export")

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    create_input_parameters_sheet(wb, data_volume_gb_day, message_rate, avg_message_size_kb,
                                   retention_days, partitions, replication_factor, peak_avg_ratio)

    create_calculated_metrics_sheet(wb, data_volume_gb_day, message_rate, avg_message_size_kb,
                                     retention_days, replication_factor, peak_avg_ratio)

    create_cost_breakdown_sheet(wb, costs, storage_cost_per_gb_month, throughput_cost_per_mbps_month,
                                 network_cost_per_gb_month, partition_cost_per_partition_month,
                                 retention_overhead_pct, data_volume_gb_day, message_rate,
                                 retention_days, replication_factor, partitions)

    create_cost_drivers_sheet(wb, data_volume_gb_day, message_rate, retention_days,
                               partitions, replication_factor, peak_avg_ratio)

    create_methodology_sheet(wb, data_volume_gb_day, message_rate, retention_days,
                              replication_factor, partitions, avg_message_size_kb, peak_avg_ratio,
                              storage_cost_per_gb_month, throughput_cost_per_mbps_month,
                              network_cost_per_gb_month, partition_cost_per_partition_month)

    create_pricing_sheet(wb, storage_cost_per_gb_month, throughput_cost_per_mbps_month,
                          network_cost_per_gb_month, partition_cost_per_partition_month,
                          retention_overhead_pct)

    create_optimization_sheet(wb)

    create_assumptions_sheet(wb)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def apply_header_style(cell, fill_color="333333"):
    """Apply consistent header styling"""
    cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )


def apply_cell_style(cell, is_header=False, is_bold=False, bg_color=None):
    """Apply consistent cell styling"""
    if is_header:
        apply_header_style(cell)
    else:
        cell.font = Font(name='Calibri', size=11, bold=is_bold)
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )


def create_input_parameters_sheet(wb, data_volume_gb_day, message_rate, avg_message_size_kb,
                                   retention_days, partitions, replication_factor, peak_avg_ratio):
    """Create INPUT PARAMETERS sheet"""
    ws = wb.create_sheet("Input Parameters")

    # Title
    ws.merge_cells('A1:C1')
    title = ws['A1']
    title.value = 'Confluent Technical Cost Model Analysis'
    title.font = Font(name='Calibri', size=14, bold=True)
    title.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    subtitle = ws['A2']
    subtitle.value = 'Infrastructure Capacity Planning & Cost Estimation'
    subtitle.font = Font(name='Calibri', size=11, italic=True)
    subtitle.alignment = Alignment(horizontal='center')

    # Headers
    ws['A4'] = 'INPUT PARAMETERS'
    ws['A4'].font = Font(bold=True, size=12)

    ws['A5'] = 'Parameter'
    ws['B5'] = 'Value'
    ws['C5'] = 'Unit'
    apply_header_style(ws['A5'])
    apply_header_style(ws['B5'])
    apply_header_style(ws['C5'])

    # Data rows
    params = [
        ('Data Volume', data_volume_gb_day, 'GB/day'),
        ('Message Rate', message_rate, 'messages/sec'),
        ('Average Message Size', avg_message_size_kb, 'KB'),
        ('Retention Period', retention_days, 'days'),
        ('Partitions', partitions, 'count'),
        ('Replication Factor', replication_factor, 'replicas'),
        ('Peak to Average Ratio', peak_avg_ratio, 'x')
    ]

    row = 6
    for param, value, unit in params:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        ws[f'C{row}'] = unit
        apply_cell_style(ws[f'A{row}'])
        apply_cell_style(ws[f'B{row}'])
        apply_cell_style(ws[f'C{row}'])
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20


def create_calculated_metrics_sheet(wb, data_volume_gb_day, message_rate, avg_message_size_kb,
                                     retention_days, replication_factor, peak_avg_ratio):
    """Create CALCULATED METRICS sheet"""
    ws = wb.create_sheet("Calculated Metrics")

    # Calculate metrics
    total_storage_gb = data_volume_gb_day * retention_days * replication_factor
    avg_throughput_mb = (message_rate * avg_message_size_kb) / 1024
    peak_throughput_mb = avg_throughput_mb * peak_avg_ratio
    monthly_data_transfer = data_volume_gb_day * 30

    # Headers
    ws['A1'] = 'CALCULATED METRICS'
    ws['A1'].font = Font(bold=True, size=12)

    ws['A2'] = 'Metric'
    ws['B2'] = 'Value'
    ws['C2'] = 'Unit'
    apply_header_style(ws['A2'])
    apply_header_style(ws['B2'])
    apply_header_style(ws['C2'])

    # Data
    metrics = [
        ('Total Storage Required', round(total_storage_gb, 2), 'GB'),
        ('Average Throughput', round(avg_throughput_mb, 2), 'MB/s'),
        ('Peak Throughput', round(peak_throughput_mb, 2), 'MB/s'),
        ('Monthly Data Transfer', round(monthly_data_transfer, 2), 'GB')
    ]

    row = 3
    for metric, value, unit in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'C{row}'] = unit
        apply_cell_style(ws[f'A{row}'])
        apply_cell_style(ws[f'B{row}'])
        apply_cell_style(ws[f'C{row}'])
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15


def create_cost_breakdown_sheet(wb, costs, storage_cost, throughput_cost, network_cost,
                                 partition_cost, retention_overhead, data_volume_gb_day,
                                 message_rate, retention_days, replication_factor, partitions):
    """Create ANNUAL COST BREAKDOWN sheet with formulas"""
    ws = wb.create_sheet("Cost Breakdown")

    # Headers
    ws['A1'] = 'ANNUAL COST BREAKDOWN'
    ws['A1'].font = Font(bold=True, size=12)

    ws['A2'] = 'Component'
    ws['B2'] = 'Annual Cost'
    ws['C2'] = 'Calculation'
    apply_header_style(ws['A2'])
    apply_header_style(ws['B2'])
    apply_header_style(ws['C2'], fill_color="4472C4")

    # Calculate individual costs
    storage_annual = (data_volume_gb_day * retention_days * replication_factor * storage_cost * 12)
    throughput_annual = ((message_rate * 1) / 1024 * throughput_cost * 12)  # Assuming 1KB avg
    network_annual = (data_volume_gb_day * 30 * network_cost * 12)
    partition_annual = (partitions * partition_cost * 12)
    retention_cost = storage_annual * (retention_overhead / 100)

    # Data rows with formulas as text
    cost_items = [
        ('Storage', storage_annual,
         f'{data_volume_gb_day:.0f} GB/day × {retention_days} days × {replication_factor} replicas × ${storage_cost}/GB × 12'),
        ('Throughput', throughput_annual,
         f'{message_rate:,.0f} msg/s × 1 KB/msg × {throughput_cost:.2f} $/MB/s × 12'),
        ('Network', network_annual,
         f'{data_volume_gb_day:.0f} GB/day × 30 days × ${network_cost}/GB × 12'),
        ('Partitions', partition_annual,
         f'{partitions} partitions × ${partition_cost}/partition × 12'),
        ('Retention', retention_cost,
         f'Additional {retention_overhead}% for retention overhead')
    ]

    row = 3
    for component, annual, calc in cost_items:
        ws[f'A{row}'] = component
        ws[f'B{row}'] = round(annual, 0)
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'] = calc
        apply_cell_style(ws[f'A{row}'])
        apply_cell_style(ws[f'B{row}'])
        apply_cell_style(ws[f'C{row}'])
        row += 1

    # Total row
    total = sum([item[1] for item in cost_items])
    ws[f'A{row}'] = 'TOTAL'
    ws[f'B{row}'] = round(total, 0)
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = f'Monthly: ${round(total/12, 0):,}'
    apply_cell_style(ws[f'A{row}'], is_bold=True, bg_color="D9E9F7")
    apply_cell_style(ws[f'B{row}'], is_bold=True, bg_color="D9E9F7")
    apply_cell_style(ws[f'C{row}'], is_bold=True, bg_color="D9E9F7")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 70


def create_cost_drivers_sheet(wb, data_volume_gb_day, message_rate, retention_days,
                                partitions, replication_factor, peak_avg_ratio):
    """Create COST SENSITIVITY ANALYSIS sheet"""
    ws = wb.create_sheet("Cost Drivers")

    ws['A1'] = 'COST SENSITIVITY ANALYSIS'
    ws['A1'].font = Font(bold=True, size=12)

    ws['A2'] = 'Driver'
    ws['B2'] = 'Impact Level'
    ws['C2'] = 'Notes'
    apply_header_style(ws['A2'])
    apply_header_style(ws['B2'])
    apply_header_style(ws['C2'])

    drivers = [
        ('Data Volume (GB/day)', 'HIGH', 'Linear relationship with storage and network costs'),
        ('Message Rate (msg/s)', 'HIGH', 'Directly impacts throughput costs'),
        ('Retention Period (days)', 'MEDIUM', 'Multiplies storage requirements'),
        ('Partitions', 'MEDIUM', 'Fixed cost per partition'),
        ('Replication Factor', 'HIGH', 'Multiplies storage (2x for RF=2, 3x for RF=3)'),
        ('Peak to Average Ratio', 'MEDIUM', 'Affects capacity planning but not base cost')
    ]

    row = 3
    for driver, impact, notes in drivers:
        ws[f'A{row}'] = driver
        ws[f'B{row}'] = impact
        ws[f'C{row}'] = notes

        # Color code impact level
        bg_color = "FFC7CE" if impact == "HIGH" else ("FFEB9C" if impact == "MEDIUM" else "C6EFCE")
        apply_cell_style(ws[f'A{row}'])
        apply_cell_style(ws[f'B{row}'], is_bold=True, bg_color=bg_color)
        apply_cell_style(ws[f'C{row}'])
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60


def create_methodology_sheet(wb, data_volume_gb_day, message_rate, retention_days,
                               replication_factor, partitions, avg_message_size_kb, peak_avg_ratio,
                               storage_cost, throughput_cost, network_cost, partition_cost):
    """Create METHODOLOGY DETAILS sheet"""
    ws = wb.create_sheet("Methodology")

    ws['A1'] = 'METHODOLOGY DETAILS'
    ws['A1'].font = Font(bold=True, size=12)

    ws['A3'] = 'Calculation'
    ws['B3'] = 'Formula'
    ws['C3'] = 'Example Values'
    apply_header_style(ws['A3'])
    apply_header_style(ws['B3'])
    apply_header_style(ws['C3'])

    methodology = [
        ('Storage Calculation',
         'GB/day × Retention Days × Replication Factor',
         f'{data_volume_gb_day} GB/day × {retention_days} days × {replication_factor} replicas = {data_volume_gb_day * retention_days * replication_factor:,.0f} GB'),
        ('Throughput Calculation',
         'msg/s × message_size × peak_ratio',
         f'{message_rate:,.0f} msg/s × {avg_message_size_kb} KB/msg × {peak_avg_ratio}x avg = {message_rate * avg_message_size_kb * peak_avg_ratio / 1024:.2f} MB/s avg, {message_rate * avg_message_size_kb * peak_avg_ratio:.2f} MB/s peak'),
        ('Network Calculation',
         'GB/day × 30 days × rate/GB',
         f'{data_volume_gb_day} GB/day × 30 days × ${network_cost}/GB = ${data_volume_gb_day * 30 * network_cost:,.0f}/month'),
        ('Partition Cost',
         'Partitions × rate/partition',
         f'{partitions} partitions × ${partition_cost}/partition/month')
    ]

    row = 4
    for calc, formula, example in methodology:
        ws[f'A{row}'] = calc
        ws[f'B{row}'] = formula
        ws[f'C{row}'] = example
        apply_cell_style(ws[f'A{row}'], is_bold=True)
        apply_cell_style(ws[f'B{row}'])
        apply_cell_style(ws[f'C{row}'])
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 70


def create_pricing_sheet(wb, storage_cost, throughput_cost, network_cost,
                          partition_cost, retention_overhead):
    """Create PRICING ASSUMPTIONS sheet"""
    ws = wb.create_sheet("Pricing")

    ws['A1'] = 'PRICING ASSUMPTIONS'
    ws['A1'].font = Font(bold=True, size=12)

    ws['A2'] = 'Component'
    ws['B2'] = 'Unit Rate (Monthly)'
    ws['C2'] = 'Unit Rate (Annual)'
    apply_header_style(ws['A2'])
    apply_header_style(ws['B2'])
    apply_header_style(ws['C2'])

    pricing = [
        ('Storage', f'${storage_cost:.2f}/GB', f'${storage_cost * 12:.2f}/GB'),
        ('Throughput', f'${throughput_cost:.2f}/MBps', f'${throughput_cost * 12:.2f}/MBps'),
        ('Network Transfer', f'${network_cost:.2f}/GB', f'${network_cost * 12:.2f}/GB'),
        ('Partitions', f'${partition_cost:.2f}/partition', f'${partition_cost * 12:.2f}/partition'),
        ('Retention Management', f'{retention_overhead}% of storage', f'{retention_overhead}% of storage')
    ]

    row = 3
    for component, monthly, annual in pricing:
        ws[f'A{row}'] = component
        ws[f'B{row}'] = monthly
        ws[f'C{row}'] = annual
        apply_cell_style(ws[f'A{row}'])
        apply_cell_style(ws[f'B{row}'])
        apply_cell_style(ws[f'C{row}'])
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25


def create_optimization_sheet(wb):
    """Create COST OPTIMIZATION RECOMMENDATIONS sheet"""
    ws = wb.create_sheet("Optimization")

    ws['A1'] = 'COST OPTIMIZATION RECOMMENDATIONS'
    ws['A1'].font = Font(bold=True, size=12)

    ws['A2'] = 'Category'
    ws['B2'] = 'Recommendation'
    ws['C2'] = 'Potential Savings'
    apply_header_style(ws['A2'])
    apply_header_style(ws['B2'])
    apply_header_style(ws['C2'])

    recommendations = [
        ('Storage', 'Implement tiered storage for data older than 48 hours', '20-40% storage cost reduction'),
        ('Throughput', 'Enable message batching and compression', '15-25% throughput cost reduction'),
        ('Retention', 'Audit and reduce retention periods where possible', 'Direct 1:1 reduction'),
        ('Partitions', 'Right-size partition count based on actual parallelism', 'Varies by over-provisioning'),
        ('Network', 'Optimize message payloads and enable compression', '10-30% network cost reduction'),
        ('Replication', 'Evaluate if 3x replication is required for all data', '33% storage cost reduction if reduced to 2x'),
        ('Peak Ratio', 'Implement auto-scaling to handle peaks more efficiently', '10-20% throughput cost reduction')
    ]

    row = 3
    for category, recommendation, savings in recommendations:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = recommendation
        ws[f'C{row}'] = savings
        apply_cell_style(ws[f'A{row}'], is_bold=True)
        apply_cell_style(ws[f'B{row}'])
        apply_cell_style(ws[f'C{row}'], bg_color="C6EFCE")
        row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 30


def create_assumptions_sheet(wb):
    """Create ASSUMPTIONS & CONSTRAINTS sheet"""
    ws = wb.create_sheet("Assumptions")

    ws['A1'] = 'ASSUMPTIONS & CONSTRAINTS'
    ws['A1'].font = Font(bold=True, size=12)

    assumptions = [
        '1. Pricing based on representative Kafka/Confluent cloud infrastructure costs',
        '2. Actual costs may vary based on specific cloud provider (AWS/Azure/GCP) and region',
        '3. Does not include additional costs for: Schema Registry, Connect clusters, ksqlDB, Enterprise support',
        '4. Network costs assume standard egress rates; ingress typically free',
        '5. Peak to average ratio assumes consistent traffic patterns; may need adjustment for bursty workloads',
        '6. Retention management overhead estimated at 5%; actual may vary',
        '7. Partition costs include compute/memory overhead for partition leadership and replication',
        '8. Storage costs include overhead for indexing, compression, and operational buffers'
    ]

    row = 3
    for assumption in assumptions:
        ws[f'A{row}'] = assumption
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    ws.column_dimensions['A'].width = 120

    # Add generation timestamp
    ws[f'A{row + 2}'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws[f'A{row + 2}'].font = Font(italic=True, size=9)

    ws[f'A{row + 3}'] = 'Model Version: Technical Cost Model v1.0'
    ws[f'A{row + 3}'].font = Font(italic=True, size=9)

    ws[f'A{row + 4}'] = 'Contact: Infrastructure Planning Team for questions or clarifications'
    ws[f'A{row + 4}'].font = Font(italic=True, size=9)
