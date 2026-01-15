import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def format_in_thousands(value):
    return f"${value:,.0f}"

def calculate_rom_costs(config):
    # Calculate based on new feed_configs structure
    num_ingests = config.get('num_ingests', 1)
    feed_configs = config.get('feed_configs', [{'inbound': 1, 'outbound': 1, 'partitions': 0.048}])
    records_per_day = config.get('records_per_day', 5000)

    # Calculate total feeds and partitions from feed_configs
    total_inbound_feeds = sum(f['inbound'] for f in feed_configs)
    total_outbound_feeds = sum(f['outbound'] for f in feed_configs)
    total_feeds = num_ingests  # Number of separate ingests
    total_partitions = sum(f['partitions'] for f in feed_configs)

    # Calculate engineering costs (scales with number of topics)
    inbound_cost = total_inbound_feeds * config['inbound_hours'] * config['de_hourly_rate']
    outbound_cost = total_outbound_feeds * config['outbound_hours'] * config['de_hourly_rate']
    normalization_cost = config['normalization_hours'] * config['de_hourly_rate'] * total_feeds
    workspace_setup = config['workspace_setup_cost']

    one_time_development = inbound_cost + outbound_cost + normalization_cost + workspace_setup

    # Cloud costs - scale with partition usage
    # Base costs from config
    base_confluent_annual = config['confluent_annual_cost']
    base_gcp_annual = config['gcp_per_feed_annual_cost']

    # Network capacity: Reference shows 100 total partitions across all sources
    # Partition ratio determines resource utilization
    TOTAL_NETWORK_PARTITIONS = 100.0
    partition_utilization = total_partitions / TOTAL_NETWORK_PARTITIONS

    # Confluent cost scales with partitions (more partitions = more throughput)
    confluent_cost = base_confluent_annual * total_feeds * (1 + partition_utilization)

    # GCP cost scales with both feeds and partitions
    # Storage scales with records per day (rough estimate: 1KB per record)
    records_per_year = records_per_day * 365
    storage_gb_per_year = records_per_year / (1024 * 1024)  # Convert to GB
    storage_multiplier = 1 + (storage_gb_per_year / 1000)  # Scale factor

    gcp_cost = base_gcp_annual * total_feeds * storage_multiplier

    # Network costs based on partition usage
    # Base network cost (from existing flat costs in app)
    base_network_annual = 120000  # $10k/month baseline
    network_cost = base_network_annual * partition_utilization

    first_year_cloud_cost = confluent_cost + gcp_cost + network_cost

    initial_investment = [{
        'year': config['start_year'],
        'data_engineering': one_time_development,
        'cloud_infrastructure': first_year_cloud_cost,
        'total': one_time_development + first_year_cloud_cost
    }]

    operating_variance = []
    cloud_infrastructure_7year = first_year_cloud_cost
    operating_variance_6year = 0

    for i in range(1, 7):
        year = config['start_year'] + i
        escalated_cloud_cost = first_year_cloud_cost * ((1 + config['escalation_rate']) ** i)
        cloud_infrastructure_7year += escalated_cloud_cost
        operating_variance_6year += escalated_cloud_cost

        operating_variance.append({
            'year': year,
            'data_engineering': 0,
            'cloud_infrastructure': escalated_cloud_cost,
            'total': escalated_cloud_cost
        })

    total_project_cost = one_time_development + cloud_infrastructure_7year

    return {
        'initial_investment': initial_investment,
        'operating_variance': operating_variance,
        'total_feeds': total_feeds,
        'total_partitions': total_partitions,
        'total_inbound_feeds': total_inbound_feeds,
        'total_outbound_feeds': total_outbound_feeds,
        'feed_configs': feed_configs,
        'records_per_day': records_per_day,
        'partition_utilization_pct': partition_utilization * 100,
        'breakdown': {
            'inbound_cost': inbound_cost,
            'outbound_cost': outbound_cost,
            'normalization_cost': normalization_cost,
            'workspace_setup': workspace_setup,
            'confluent_cost': confluent_cost,
            'gcp_cost': gcp_cost,
            'network_cost': network_cost,
            'one_time_development': one_time_development,
            'cloud_infrastructure_7year': cloud_infrastructure_7year,
            'operating_variance_6year': operating_variance_6year,
            'total_project_cost': total_project_cost,
            'first_year_cloud_cost': first_year_cloud_cost
        }
    }

def generate_rom_export(config):
    """Generate CSV format for backward compatibility"""
    results = calculate_rom_costs(config)
    lines = []

    lines.append('Confluent Feed ROM - Rough Order of Magnitude')
    lines.append('')

    years = [config['start_year'] + i for i in range(12)]
    lines.append('Fiscal Year,' + ','.join(map(str, years)) + ',Total')

    lines.append('INITIAL INVESTMENT EXPENSE')
    initial_de = results['initial_investment'][0]['data_engineering']
    de_line = f"Data Engineering,{format_in_thousands(initial_de)}" + ',,,,,,,,,,,' + f",{format_in_thousands(initial_de)}"
    lines.append(de_line)

    lines.append('Data Strategy and Governance,,,,,,,,,,,,$-')
    lines.append('Enterprise Reporting and Dashboard,,,,,,,,,,,,$-')
    lines.append('Advance Modeling,,,,,,,,,,,,$-')
    lines.append('Service Performance,,,,,,,,,,,,$-')

    initial_cloud = results['initial_investment'][0]['cloud_infrastructure']
    cloud_costs = [format_in_thousands(ov['cloud_infrastructure']) for ov in results['operating_variance']]
    lines.append(
        f"GCP/GKE/Confluent,{format_in_thousands(initial_cloud)},{','.join(cloud_costs)},,,,,,{format_in_thousands(results['breakdown']['cloud_infrastructure_7year'])}"
    )

    initial_total = results['initial_investment'][0]['total']
    total_line = f"TOTAL,{format_in_thousands(initial_total)},{','.join(cloud_costs)},,,,,,{format_in_thousands(results['breakdown']['total_project_cost'])}"
    lines.append(total_line)

    lines.append('')
    lines.append('')

    lines.append('Fiscal Year,' + ','.join(map(str, years)) + ',Total')
    lines.append('OPERATING VARIANCE')

    op_var_costs = [format_in_thousands(ov['cloud_infrastructure']) for ov in results['operating_variance']]
    lines.append(f"Data Engineering,,{','.join(op_var_costs)},,,,,,{format_in_thousands(results['breakdown']['operating_variance_6year'])}")

    lines.append('Data Strategy and Governance,,,,,,,,,,,,$-')
    lines.append('Enterprise Reporting and Dashboard,,,,,,,,,,,,$-')
    lines.append('Advance Modeling,,,,,,,,,,,,$-')
    lines.append('Service Performance,,,,,,,,,,,,$-')

    lines.append(f"TOTAL,,{','.join(op_var_costs)},,,,,,{format_in_thousands(results['breakdown']['operating_variance_6year'])}")

    lines.append('')
    lines.append('')

    lines.append('Summary')
    lines.append('Capital,$-')
    lines.append(f"Expense,{format_in_thousands(results['breakdown']['total_project_cost'])}")
    lines.append(f"Variance,{format_in_thousands(results['breakdown']['operating_variance_6year'])}")
    lines.append(f"Total,{format_in_thousands(results['breakdown']['total_project_cost'])}")

    lines.append('')
    lines.append('')
    lines.append(f"Escalation Rate,{config['escalation_rate'] * 100:.1f}%")

    lines.append('')
    lines.append('Note*')
    lines.append('"Estimate based on latest Payroll 2.0 scaling factors"')
    lines.append('"ROM may require revision as detailed requirements are finalized"')

    lines.append('')
    lines.append('Assumptions:')
    lines.append(f"1,ROM covers {results['total_feeds']} EEB ingest feed(s) with inbound/outbound data processing capabilities")
    lines.append('2,Feed ingests data with complex processing requirements')
    lines.append('3,Includes event data with facility impacts and workflow approvals')
    lines.append('4,Feed includes data normalization and standardization requirements')
    lines.append('5,Workspace/Environment setup costs included')
    lines.append(f"6,Confluent platform required for real-time streaming: {format_in_thousands(config['confluent_annual_cost'])} per feed per year")
    lines.append(f"7,GCP/GKE infrastructure cost: {format_in_thousands(config['gcp_per_feed_annual_cost'])} per feed per year for compute and storage")
    lines.append('8,ROM based on current understanding of high level requirements & known attributes')
    lines.append('9,As requirements are refined/finalized the ROM may need to be revised')

    lines.append('')
    lines.append('Timeline')
    lines.append(f"FY{config['start_year']}-FY{config['start_year'] + 6}")
    lines.append(f"12,FY{config['start_year']}: {format_in_thousands(initial_total)} (Data Engineering + Cloud infrastructure setup - starting in 3 weeks)")
    lines.append(f"13,FY{config['start_year'] + 1}-{config['start_year'] + 6}: {format_in_thousands(results['breakdown']['operating_variance_6year'] / 6)} annually (ongoing cloud operations with {config['escalation_rate'] * 100:.1f}% escalation) plus Operating Variance")

    lines.append('')
    lines.append('Cost Breakdown per Feed:')
    lines.append(f"14,Create inbound ingest: {format_in_thousands(config['inbound_hours'] * config['de_hourly_rate'])},{round(config['inbound_hours'])} ({round(config['inbound_hours'])} hours)")
    lines.append(f"15,Create outbound enterprise data assets: {format_in_thousands(config['outbound_hours'] * config['de_hourly_rate'])},{round(config['outbound_hours'])} ({round(config['outbound_hours'])} hours)")
    lines.append(f"16,Data normalization and standardization: {format_in_thousands(results['breakdown']['normalization_cost'])},{round(config['normalization_hours'])} ({config['normalization_hours']} hours - {results['total_feeds']} feeds)")
    lines.append(f"17,Workspace/Environment/Subscription Prep: {format_in_thousands(config['workspace_setup_cost'])}")
    lines.append(f"18,Annual Confluent platform cost: {format_in_thousands(config['confluent_annual_cost'])},{round(config['confluent_annual_cost'])}")
    lines.append(f"19,Annual GCP/GKE cost: {format_in_thousands(config['gcp_per_feed_annual_cost'])},{round(config['gcp_per_feed_annual_cost'])} per feed")

    lines.append('')
    lines.append(f"Total {results['total_feeds']}-Feed Investment")
    lines.append(f"{results['total_feeds']}-Feed Investment")
    lines.append(f"21,Data Engineering: {format_in_thousands(results['breakdown']['one_time_development'])},{round(results['breakdown']['one_time_development'] / 1000)} (one-time development)")
    lines.append(f"22,Cloud Infrastructure: {format_in_thousands(results['breakdown']['cloud_infrastructure_7year'])},{round(results['breakdown']['cloud_infrastructure_7year'] / 1000)} (7-year operational costs with {config['escalation_rate'] * 100:.1f}% escalation)")
    lines.append(f"23,Operating Variance: {format_in_thousands(results['breakdown']['operating_variance_6year'])},{round(results['breakdown']['operating_variance_6year'] / 1000)} (6-year escalated costs)")
    lines.append(f"24,Total Project Cost: {format_in_thousands(results['breakdown']['total_project_cost'])},{round(results['breakdown']['total_project_cost'] / 1000)}")

    return '\n'.join(lines)


def generate_rom_export_excel_de_only(config):
    """Generate Data Engineering Only ROM"""
    if not HAS_OPENPYXL:
        return generate_rom_export(config).encode('utf-8')

    results = calculate_rom_costs(config)
    wb = Workbook()
    ws = wb.active
    ws.title = "ROM - DE Only"

    # Color scheme - USPS Blue
    usps_blue = "004B87"
    header_fill = PatternFill(start_color=usps_blue, end_color=usps_blue, fill_type="solid")
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    title_font = Font(name='Calibri', size=16, bold=True, color=usps_blue)
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws.merge_cells(f'A{row}:E{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = 'Confluent Feed ROM - Data Engineering Only'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Feed Configuration Summary
    ws[f'A{row}'] = 'Feed Configuration Summary'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    ws.cell(row, 1, f"Number of Ingests: {results['total_feeds']}").font = normal_font
    row += 1
    ws.cell(row, 1, f"Total Inbound Topics: {results['total_inbound_feeds']}").font = normal_font
    row += 1
    ws.cell(row, 1, f"Total Outbound Topics: {results['total_outbound_feeds']}").font = normal_font
    row += 2

    # Cost Breakdown
    ws[f'A{row}'] = 'DATA ENGINEERING COSTS (One-Time)'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = yellow_fill
    row += 1

    de_items = [
        ('Inbound Development', results['breakdown']['inbound_cost']),
        ('Outbound Development', results['breakdown']['outbound_cost']),
        ('Normalization', results['breakdown']['normalization_cost']),
        ('Workspace Setup', results['breakdown']['workspace_setup']),
    ]

    for label, value in de_items:
        ws.cell(row, 1, label).border = thin_border
        ws.cell(row, 2, value).number_format = '$#,##0'
        ws.cell(row, 2).border = thin_border
        row += 1

    ws.cell(row, 1, 'TOTAL').font = bold_font
    ws.cell(row, 1).border = thin_border
    ws.cell(row, 1).fill = header_fill
    ws.cell(row, 1).font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    ws.cell(row, 2, results['breakdown']['one_time_development']).number_format = '$#,##0'
    ws.cell(row, 2).border = thin_border
    ws.cell(row, 2).font = Font(name='Calibri', size=12, bold=True, color="FF0000")
    row += 2

    # Assumptions
    ws[f'A{row}'] = 'Assumptions:'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    assumptions = [
        f"ROM covers {results['total_feeds']} EEB ingest feed(s) with inbound/outbound data processing capabilities",
        f"Total {results['total_inbound_feeds']} inbound topics and {results['total_outbound_feeds']} outbound topics",
        "Feed ingests data with complex processing requirements",
        "Includes event data with facility impacts and workflow approvals",
        "Feed includes data normalization and standardization requirements",
        "Workspace/Environment setup costs included",
        f"Hourly rate: {format_in_thousands(config['de_hourly_rate'])}/hour",
        f"Inbound hours per topic: {config['inbound_hours']:.1f} hours",
        f"Outbound hours per topic: {config['outbound_hours']:.1f} hours"
    ]

    for i, assumption in enumerate(assumptions, start=1):
        ws.cell(row, 1, f"{i}. {assumption}").font = normal_font
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_rom_export_excel_cloud_only(config):
    """Generate Cloud Infrastructure Only ROM"""
    if not HAS_OPENPYXL:
        return generate_rom_export(config).encode('utf-8')

    results = calculate_rom_costs(config)
    wb = Workbook()
    ws = wb.active
    ws.title = "ROM - Cloud Only"

    # Color scheme - USPS Blue
    usps_blue = "004B87"
    header_fill = PatternFill(start_color=usps_blue, end_color=usps_blue, fill_type="solid")
    light_blue = PatternFill(start_color="D9E9F7", end_color="D9E9F7", fill_type="solid")
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    title_font = Font(name='Calibri', size=16, bold=True, color=usps_blue)
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws.merge_cells(f'A{row}:N{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = 'Confluent Feed ROM - Cloud Infrastructure Only'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Feed Configuration Summary
    ws[f'A{row}'] = 'Feed Configuration Summary'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    ws.cell(row, 1, f"Number of Ingests: {results['total_feeds']}").font = normal_font
    row += 1
    ws.cell(row, 1, f"Total Partitions: {results['total_partitions']:.3f}").font = normal_font
    row += 1
    ws.cell(row, 1, f"Network Utilization: {results['partition_utilization_pct']:.2f}%").font = normal_font
    row += 1
    ws.cell(row, 1, f"Records per Day: {results['records_per_day']:,}").font = normal_font
    row += 2

    # Annual Cost Breakdown
    ws[f'A{row}'] = 'CLOUD INFRASTRUCTURE (Annual)'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_blue
    row += 1

    ws.cell(row, 1, 'Confluent Cost').border = thin_border
    ws.cell(row, 2, results['breakdown']['confluent_cost']).number_format = '$#,##0'
    ws.cell(row, 2).border = thin_border
    row += 1

    ws.cell(row, 1, 'GCP Cost').border = thin_border
    ws.cell(row, 2, results['breakdown']['gcp_cost']).number_format = '$#,##0'
    ws.cell(row, 2).border = thin_border
    row += 1

    ws.cell(row, 1, 'Network Cost').border = thin_border
    ws.cell(row, 2, results['breakdown']['network_cost']).number_format = '$#,##0'
    ws.cell(row, 2).border = thin_border
    row += 1

    ws.cell(row, 1, 'First Year Total').font = bold_font
    ws.cell(row, 1).border = thin_border
    ws.cell(row, 1).fill = header_fill
    ws.cell(row, 1).font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    ws.cell(row, 2, results['breakdown']['first_year_cloud_cost']).number_format = '$#,##0'
    ws.cell(row, 2).border = thin_border
    ws.cell(row, 2).font = Font(name='Calibri', size=12, bold=True, color="FF0000")
    row += 2

    # 7-Year Projection
    years = [config['start_year'] + i for i in range(12)]
    ws[f'A{row}'] = 'Fiscal Year'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].border = thin_border

    for idx, year in enumerate(years, start=2):
        cell = ws.cell(row, idx, year)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ws.cell(row, 14, 'Total').font = header_font
    ws.cell(row, 14).fill = header_fill
    ws.cell(row, 14).alignment = Alignment(horizontal='center')
    ws.cell(row, 14).border = thin_border
    row += 1

    # GCP/GKE/Confluent row
    initial_cloud = results['initial_investment'][0]['cloud_infrastructure']
    ws.cell(row, 1, 'GCP/GKE/Confluent').border = thin_border
    ws.cell(row, 2, initial_cloud).number_format = '$#,##0'
    ws.cell(row, 2).border = thin_border

    for idx, ov in enumerate(results['operating_variance'], start=3):
        ws.cell(row, idx, ov['cloud_infrastructure']).number_format = '$#,##0'
        ws.cell(row, idx).border = thin_border

    ws.cell(row, 14, results['breakdown']['cloud_infrastructure_7year']).number_format = '$#,##0'
    ws.cell(row, 14).border = thin_border
    ws.cell(row, 1).fill = light_blue
    row += 2

    # Escalation Rate
    ws.cell(row, 1, 'Escalation Rate:').font = bold_font
    ws.cell(row, 2, f"{config['escalation_rate'] * 100:.1f}%")
    row += 2

    # Assumptions
    ws[f'A{row}'] = 'Assumptions:'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    assumptions = [
        f"ROM covers {results['total_feeds']} EEB ingest feed(s)",
        f"Network utilization: {results['partition_utilization_pct']:.2f}% ({results['total_partitions']:.3f} partitions out of 100 total)",
        f"Daily volume: {results['records_per_day']:,} records per day",
        f"Confluent platform required for real-time streaming: {format_in_thousands(config['confluent_annual_cost'])} base cost per feed per year",
        f"GCP/GKE infrastructure cost: {format_in_thousands(config['gcp_per_feed_annual_cost'])} base cost per feed per year",
        f"Network costs: {format_in_thousands(120000)} baseline, scaled by partition utilization",
        f"Escalation rate: {config['escalation_rate'] * 100:.1f}% annually for years 2-7",
        "Costs scale with partition usage and data volume",
        "ROM based on current understanding of high level requirements & known attributes"
    ]

    for i, assumption in enumerate(assumptions, start=1):
        ws.cell(row, 1, f"{i}. {assumption}").font = normal_font
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 100
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_rom_export_excel(config, logo_path=None):
    """Generate formatted Excel file with complete ROM (DE + Cloud)"""
    if not HAS_OPENPYXL:
        return generate_rom_export(config).encode('utf-8')

    results = calculate_rom_costs(config)
    wb = Workbook()
    ws = wb.active
    ws.title = "ROM - Complete"

    # Color scheme - USPS Blue
    usps_blue = "004B87"
    header_fill = PatternFill(start_color=usps_blue, end_color=usps_blue, fill_type="solid")
    light_blue = PatternFill(start_color="D9E9F7", end_color="D9E9F7", fill_type="solid")
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    title_font = Font(name='Calibri', size=16, bold=True, color=usps_blue)
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws.merge_cells(f'A{row}:N{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = 'Confluent Feed ROM - Rough Order of Magnitude'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # Feed Configuration Summary
    ws[f'A{row}'] = 'Feed Configuration Summary'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    ws.cell(row, 1, f"Number of Ingests: {results['total_feeds']}").font = normal_font
    row += 1
    ws.cell(row, 1, f"Total Inbound Topics: {results['total_inbound_feeds']}").font = normal_font
    row += 1
    ws.cell(row, 1, f"Total Outbound Topics: {results['total_outbound_feeds']}").font = normal_font
    row += 1
    ws.cell(row, 1, f"Total Partitions: {results['total_partitions']:.3f}").font = normal_font
    row += 1
    ws.cell(row, 1, f"Network Utilization: {results['partition_utilization_pct']:.2f}%").font = normal_font
    row += 1
    ws.cell(row, 1, f"Records per Day: {results['records_per_day']:,}").font = normal_font
    row += 1

    # Feed Patterns
    ws[f'A{row}'] = 'Feed Patterns:'
    ws[f'A{row}'].font = bold_font
    row += 1

    for i, feed in enumerate(results['feed_configs'], start=1):
        pattern_text = f"  Feed {i}: {feed['inbound']} inbound → {feed['outbound']} outbound | {feed['partitions']:.3f} partitions"
        ws.cell(row, 1, pattern_text).font = normal_font
        row += 1

    row += 1

    # Cost Breakdown Summary
    ws[f'A{row}'] = 'Cost Breakdown Summary'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    row += 1

    # Data Engineering Costs
    ws[f'A{row}'] = 'DATA ENGINEERING (One-Time):'
    ws[f'A{row}'].font = bold_font
    row += 1
    ws.cell(row, 1, f"  Inbound Development: ${results['breakdown']['inbound_cost']:,.0f}").font = normal_font
    row += 1
    ws.cell(row, 1, f"  Outbound Development: ${results['breakdown']['outbound_cost']:,.0f}").font = normal_font
    row += 1
    ws.cell(row, 1, f"  Normalization: ${results['breakdown']['normalization_cost']:,.0f}").font = normal_font
    row += 1
    ws.cell(row, 1, f"  Workspace Setup: ${results['breakdown']['workspace_setup']:,.0f}").font = normal_font
    row += 1
    ws[f'A{row}'] = f"  Total One-Time: ${results['breakdown']['one_time_development']:,.0f}"
    ws[f'A{row}'].font = bold_font
    row += 2

    # Cloud Infrastructure Costs
    ws[f'A{row}'] = 'CLOUD INFRASTRUCTURE (Annual):'
    ws[f'A{row}'].font = bold_font
    row += 1
    ws.cell(row, 1, f"  Confluent Cost: ${results['breakdown']['confluent_cost']:,.0f}").font = normal_font
    row += 1
    ws.cell(row, 1, f"  GCP Cost: ${results['breakdown']['gcp_cost']:,.0f}").font = normal_font
    row += 1
    ws.cell(row, 1, f"  Network Cost: ${results['breakdown']['network_cost']:,.0f}").font = normal_font
    row += 1
    ws[f'A{row}'] = f"  First Year Total: ${results['breakdown']['first_year_cloud_cost']:,.0f}"
    ws[f'A{row}'].font = bold_font
    row += 2

    # 7-Year Projection
    ws[f'A{row}'] = '7-YEAR PROJECTION:'
    ws[f'A{row}'].font = bold_font
    row += 1
    ws.cell(row, 1, f"  Cloud Infrastructure (7 years): ${results['breakdown']['cloud_infrastructure_7year']:,.0f}").font = normal_font
    row += 1
    ws.cell(row, 1, f"  Data Engineering (One-Time): ${results['breakdown']['one_time_development']:,.0f}").font = normal_font
    row += 1
    ws[f'A{row}'] = f"  TOTAL PROJECT COST: ${results['breakdown']['total_project_cost']:,.0f}"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color="FF0000")
    row += 2

    # Year headers
    years = [config['start_year'] + i for i in range(12)]
    ws[f'A{row}'] = 'Fiscal Year'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].border = thin_border

    for idx, year in enumerate(years, start=2):
        cell = ws.cell(row, idx, year)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ws.cell(row, 14, 'Total').font = header_font
    ws.cell(row, 14).fill = header_fill
    ws.cell(row, 14).alignment = Alignment(horizontal='center')
    ws.cell(row, 14).border = thin_border
    row += 1

    # INITIAL INVESTMENT EXPENSE
    ws[f'A{row}'] = 'INITIAL INVESTMENT EXPENSE'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = yellow_fill
    row += 1

    # Data Engineering
    initial_de = results['initial_investment'][0]['data_engineering']
    ws.cell(row, 1, 'Data Engineering').border = thin_border
    ws.cell(row, 2, initial_de).number_format = '$#,##0'
    ws.cell(row, 2).border = thin_border
    ws.cell(row, 14, initial_de).number_format = '$#,##0'
    ws.cell(row, 14).border = thin_border
    row += 1

    # Empty categories
    empty_cats = ['Data Strategy and Governance', 'Enterprise Reporting and Dashboard',
                  'Advance Modeling', 'Service Performance']
    for cat in empty_cats:
        ws.cell(row, 1, cat).border = thin_border
        ws.cell(row, 14, 0).number_format = '$#,##0'
        ws.cell(row, 14).border = thin_border
        row += 1

    # GCP/GKE/Confluent
    initial_cloud = results['initial_investment'][0]['cloud_infrastructure']
    ws.cell(row, 1, 'GCP/GKE/Confluent').border = thin_border
    ws.cell(row, 2, initial_cloud).number_format = '$#,##0'
    ws.cell(row, 2).border = thin_border

    for idx, ov in enumerate(results['operating_variance'], start=3):
        ws.cell(row, idx, ov['cloud_infrastructure']).number_format = '$#,##0'
        ws.cell(row, idx).border = thin_border

    ws.cell(row, 14, results['breakdown']['cloud_infrastructure_7year']).number_format = '$#,##0'
    ws.cell(row, 14).border = thin_border
    ws.cell(row, 1).fill = light_blue
    row += 1

    # TOTAL
    initial_total = results['initial_investment'][0]['total']
    ws.cell(row, 1, 'TOTAL').font = bold_font
    ws.cell(row, 1).border = thin_border
    ws.cell(row, 1).fill = header_fill
    ws.cell(row, 1).font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    ws.cell(row, 2, initial_total).number_format = '$#,##0'
    ws.cell(row, 2).border = thin_border
    ws.cell(row, 2).font = bold_font

    for idx, ov in enumerate(results['operating_variance'], start=3):
        ws.cell(row, idx, ov['cloud_infrastructure']).number_format = '$#,##0'
        ws.cell(row, idx).border = thin_border
        ws.cell(row, idx).font = bold_font

    ws.cell(row, 14, results['breakdown']['total_project_cost']).number_format = '$#,##0'
    ws.cell(row, 14).border = thin_border
    ws.cell(row, 14).font = bold_font
    row += 2

    # Summary section
    ws[f'A{row}'] = 'Summary'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    summary_data = [
        ('Capital', 0),
        ('Expense', results['breakdown']['total_project_cost']),
        ('Variance', results['breakdown']['operating_variance_6year']),
        ('Total', results['breakdown']['total_project_cost'])
    ]

    for label, value in summary_data:
        ws.cell(row, 1, label).border = thin_border
        ws.cell(row, 1).font = bold_font if label == 'Total' else None
        ws.cell(row, 2, value).number_format = '$#,##0'
        ws.cell(row, 2).border = thin_border
        ws.cell(row, 2).font = bold_font if label == 'Total' else None
        if label == 'Total':
            ws.cell(row, 1).fill = light_blue
            ws.cell(row, 2).fill = light_blue
        row += 1

    row += 1

    # Escalation Rate
    ws.cell(row, 1, 'Escalation Rate:').font = bold_font
    ws.cell(row, 2, f"{config['escalation_rate'] * 100:.1f}%")
    row += 2

    # Notes
    ws[f'A{row}'] = 'Note*'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1
    ws.cell(row, 1, 'Estimate based on latest Payroll 2.0 scaling factors').font = normal_font
    row += 1
    ws.cell(row, 1, 'ROM may require revision as detailed requirements are finalized').font = normal_font
    row += 2

    # Assumptions
    ws[f'A{row}'] = 'Assumptions:'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    assumptions = [
        f"ROM covers {results['total_feeds']} EEB ingest feed(s) with inbound/outbound data processing capabilities",
        "Feed ingests data with complex processing requirements",
        "Includes event data with facility impacts and workflow approvals",
        "Feed includes data normalization and standardization requirements",
        "Workspace/Environment setup costs included",
        f"Confluent platform required for real-time streaming: {format_in_thousands(config['confluent_annual_cost'])} per feed per year",
        f"GCP/GKE infrastructure cost: {format_in_thousands(config['gcp_per_feed_annual_cost'])} per feed per year for compute and storage",
        "ROM based on current understanding of high level requirements & known attributes",
        "As requirements are refined/finalized the ROM may need to be revised"
    ]

    for i, assumption in enumerate(assumptions, start=1):
        ws.cell(row, 1, f"{i}. {assumption}").font = normal_font
        row += 1

    row += 1

    # Timeline
    ws[f'A{row}'] = 'Timeline'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    ws.cell(row, 1, f"FY{config['start_year']}-FY{config['start_year'] + 6}").font = bold_font
    row += 1

    initial_total = results['initial_investment'][0]['total']
    ws.cell(row, 1, f"FY{config['start_year']}: {format_in_thousands(initial_total)} (Data Engineering + Cloud infrastructure setup - starting in 3 weeks)").font = normal_font
    row += 1
    ws.cell(row, 1, f"FY{config['start_year'] + 1}-{config['start_year'] + 6}: {format_in_thousands(results['breakdown']['operating_variance_6year'] / 6)} annually (ongoing cloud operations with {config['escalation_rate'] * 100:.1f}% escalation) plus Operating Variance").font = normal_font
    row += 2

    # Cost Breakdown per Feed
    ws[f'A{row}'] = 'Cost Breakdown per Feed:'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    cost_breakdown = [
        (f"Create inbound ingest: {format_in_thousands(config['inbound_hours'] * config['de_hourly_rate'])} ({round(config['inbound_hours'])} hours)"),
        (f"Create outbound enterprise data assets: {format_in_thousands(config['outbound_hours'] * config['de_hourly_rate'])} ({round(config['outbound_hours'])} hours)"),
        (f"Data normalization and standardization: {format_in_thousands(results['breakdown']['normalization_cost'])} ({config['normalization_hours']} hours - {results['total_feeds']} feeds)"),
        (f"Workspace/Environment/Subscription Prep: {format_in_thousands(config['workspace_setup_cost'])}"),
        (f"Annual Confluent platform cost: {format_in_thousands(config['confluent_annual_cost'])}"),
        (f"Annual GCP/GKE cost: {format_in_thousands(config['gcp_per_feed_annual_cost'])} per feed")
    ]

    for item in cost_breakdown:
        ws.cell(row, 1, item).font = normal_font
        row += 1

    row += 1

    # Total Feed Investment
    ws[f'A{row}'] = f'Total {results["total_feeds"]}-Feed Investment'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = light_gray
    row += 1

    investment_items = [
        (f"Data Engineering: {format_in_thousands(results['breakdown']['one_time_development'])} (one-time development)"),
        (f"Cloud Infrastructure: {format_in_thousands(results['breakdown']['cloud_infrastructure_7year'])} (7-year operational costs with {config['escalation_rate'] * 100:.1f}% escalation)"),
        (f"Operating Variance: {format_in_thousands(results['breakdown']['operating_variance_6year'])} (6-year escalated costs)"),
        (f"Total Project Cost: {format_in_thousands(results['breakdown']['total_project_cost'])}")
    ]

    for item in investment_items:
        ws.cell(row, 1, item).font = normal_font
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 100
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
